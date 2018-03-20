#!/usr/bin/env python
# coding=utf-8

import ConfigParser
import os, time
import ZabbixApi
import WeTalkApi
from openpyxl import Workbook, load_workbook

## ZabbixWeTalkApi class begin -----------------------------------------------------------------------------------------------
class ZabbixWeTalkApi(object):
    def __init__(self, token_time=7200, msg_max_size=3000, alert_max_size=168, zbx_cfg_key='zbx_', wt_cfg_key='WX_', wt_iGet_sec='WX_iGet', wt_file_sec='WX_file', wt_iGet=False, config_file='WeiXinByGroup.conf', data_file='Alert_detail.xlsx'):
        self.WT_SPLIT_STRING = u'-------------------------------------------'
        self.WT_SPLIT_BLOCK = u'◎◎◎◎◎◎◎◎◎◎◎◎'
        self.Access_Command_dic = {
            '.gettt': u'参数: t1 t2 [zbx] [grp] [f]，t1-t2告警',
            '.help': u'参数: 无，查看所有命令帮助信息。',
            '.getnow': u'参数: [zbx]，返回指定zbx的当前告警。',
            '.getsec': u'参数: tn [zbx] [grp] [f]，秒内告警',
            '.getmin': u'参数: tn [zbx] [grp] [f]，分内告警',
            '.gethour': u'参数: tn [zbx] [grp] [f]，时内告警',
            '.getday': u'参数: tn [zbx] [grp] [f]，天内告警'
        }
        self.Time_BS_dic = {'.getsec': 1, '.getmin': 60, '.gethour': 3600, '.getday': 86400}
        self.WT_TOKEN_TIME = token_time
        self.WT_MSG_MAX_SIZE = msg_max_size
        self.ALERT_MAX_SIZE = alert_max_size
        self.zbx_cfg_key = zbx_cfg_key
        self.wt_cfg_key = wt_cfg_key
        self.wt_iGet_sec = wt_iGet_sec
        self.wt_file_sec = wt_file_sec
        self.wt_iGet = wt_iGet
        self.z_dic = {}
        self.w_dic = {}
        self.config_file = config_file
        self.data_file = data_file
        if self.config_file:
            self.ReadConfig(config_file=self.config_file)
        self.WeTalkMenu()
    
    def ReadConfig(self, config_file=''):
        # 读取配置文件信息
        z_dic = {}
        w_dic = {
            self.wt_file_sec: {'start_flag': 0, 'wapi': '', 'corp_id': '', 'corp_secret': '', 'agent_id': ''},
            self.wt_iGet_sec: {'start_flag': 0, 'wapi': '', 'corp_id': '', 'corp_secret': '', 'agent_id': ''}
        }
        #判断要读的配置文件是否存在？若不存在，返回空字典
        if not config_file:
            config_file = self.config_file
        if not os.path.exists(config_file):
            print config_file, ": config file not exist!"
            self.z_dic = z_dic
            self.w_dic = w_dic
            return z_dic, w_dic
        #读取配置文件
        cf = ConfigParser.ConfigParser()
        cf.read(config_file)
        #按配置文件的Section读取配置数据，结果存字典
        for sec in cf.sections():
            if (self.zbx_cfg_key == sec[0:len(self.zbx_cfg_key)]):
                z_dic[sec] = {}
                date_time = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(time.time()))
                try:
                    z_dic[sec]['start_flag'] = int(cf.get(sec, "start_flag").strip())
                    if z_dic[sec]['start_flag']:
                        z_dic[sec]['alert_time'] = int(cf.get(sec, "alert_time").strip())
                        z_dic[sec]['collect_time'] = int(cf.get(sec, "collect_time").strip())
                        z_dic[sec]['info_time'] = int(cf.get(sec, "info_time").strip())
                        z_dic[sec]['reset_time'] = int(cf.get(sec, "reset_time").strip())
                        z_dic[sec]['server'] = cf.get(sec, "server").strip()
                        z_dic[sec]['username'] = cf.get(sec, "username").strip()
                        z_dic[sec]['password'] = cf.get(sec, "password").strip()
                        z_dic[sec]['toparty'] = cf.get(sec, "toparty").strip()
                        z_dic[sec]['toagent'] = cf.get(sec, "toagent").replace(' ', '').split('|')
                        z_dic[sec]['alert_level'] = eval(cf.get(sec, "alert_level").strip())
                        z_dic[sec]['priority'] = cf.get(sec, "priority").strip()
                        start_time = int(time.time()) #- 60 * 60
                        z_dic[sec]['alert_last_time'] = start_time
                        z_dic[sec]['collect_last_time'] = start_time
                        z_dic[sec]['info_last_time'] = start_time
                        z_dic[sec]['reset_last_time'] = start_time
                        z_dic[sec]['alert_ws_list'] = []
                        z_dic[sec]['zbx_nick'] = cf.get(sec, "zbx_nick").strip()
                        z_dic[sec]['zbx_name'] = sec 
                        if z_dic[sec]['server'] and z_dic[sec]['username'] and z_dic[sec]['password']:
                            z_dic[sec]['zapi'] = ZabbixApi.ZabbixApi(server=z_dic[sec]['server'], username=z_dic[sec]['username'], password=z_dic[sec]['password'])
                            if not z_dic[sec]['zapi'].isLogin():
                                z_dic[sec]['zapi'] = ''
                                z_dic[sec]['start_flag'] = 0
                                print u'[ %s ] : %s 初始化【失败】！' % (date_time, sec)
                            else:
                                print u'[ %s ] : %s 初始化【成功】！' % (date_time, sec)
                        else:
                            z_dic[sec]['zapi'] = ''
                            z_dic[sec]['start_flag'] = 0
                            print u'[ %s ] : %s 初始化【失败】 url/user/pass数据缺失！' % (date_time, sec)
                    else:
                        z_dic[sec]['zapi'] = ''
                        print u'[ %s ] : %s 初始化【失败】 被手动禁用！' % (date_time, sec)
                except Exception as e:
                    z_dic[sec]['zapi'] = ''
                    z_dic[sec]['start_flag'] = 0
                    print u'[ %s ] : %s 初始化【失败】 配置项错误！' % (date_time, sec)
            elif (self.wt_iGet and (sec == self.wt_iGet_sec)) or ((not self.wt_iGet) and (sec != self.wt_iGet_sec) and (self.wt_cfg_key == sec[0:len(self.wt_cfg_key)])):
                w_dic[sec] = {}
                date_time = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(time.time()))
                try:
                    w_dic[sec]['start_flag'] = int(cf.get(sec, "start_flag").strip())
                    if w_dic[sec]['start_flag']:
                        w_dic[sec]['corp_id'] = cf.get(sec, "corp_id").strip()
                        w_dic[sec]['corp_secret'] = cf.get(sec, "corp_secret").strip()
                        w_dic[sec]['agent_id'] = cf.get(sec, "agent_id").strip()
                        if w_dic[sec]['corp_id'] and w_dic[sec]['corp_secret'] and w_dic[sec]['agent_id']:
                            w_dic[sec]['wapi'] = WeTalkApi.WeTalkApi(token_time=self.WT_TOKEN_TIME, msg_max_size=self.WT_MSG_MAX_SIZE, corp_id=w_dic[sec]['corp_id'], corp_secret=w_dic[sec]['corp_secret'], agent_id=w_dic[sec]['agent_id'])
                            date_time = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(time.time()))
                            if not w_dic[sec]['wapi'].get_access_token():
                                w_dic[sec]['wapi'] = ''
                                w_dic[sec]['start_flag'] = 0
                                print u'[ %s ] : %s 初始化【失败】！' % (date_time, sec)
                            else:
                                print u'[ %s ] : %s 初始化【成功】！' % (date_time, sec)
                        else:
                            w_dic[sec]['wapi'] = ''
                            w_dic[sec]['start_flag'] = 0
                            print u'[ %s ] : %s 初始化【失败】 url/user/pass数据缺失！' % (date_time, sec)
                    else:
                        w_dic[sec]['wapi'] = ''
                        print u'[ %s ] : %s 初始化【失败】 被手动禁用！' % (date_time, sec)
                except Exception as e:
                    w_dic[sec]['wapi'] = ''
                    w_dic[sec]['start_flag'] = 0
                    print u'[ %s ] : %s 初始化【失败】 配置项错误！' % (date_time, sec)
        self.z_dic = z_dic
        self.w_dic = w_dic
        return z_dic, w_dic

    def search_from_list(self, seList=[], ssList=[]):
        for ss in ssList:
            for se in seList:
                if se in ss:
                    return ss.replace(se, '').strip()
        return ''

    def alert_message_init(self, alert={}):
        try:
            msg = alert['message']
            msg_list = msg.replace('\r', '').split('\n')
            if len(msg_list) < 6:
                return '', {}
            rdic = {}
            rdic['group'] = ''
            rdic['host'] = alert['hosts'][0]['name']
            if not rdic['host']:
                rdic['host'] = 'Unknown'
            rdic['hostid'] = alert['hosts'][0]['hostid']
            if not rdic['hostid']:
                rdic['group'] = 'Unknown'
            rdic['IP'] = self.search_from_list(seList=[u'主机IP:', 'IP:'], ssList=msg_list)
            rdic['time'] = self.search_from_list(seList=[u'告警时间:', u'告警恢复时间:', u'time:'], ssList=msg_list)
            rdic['state'] = self.search_from_list(seList=[u'告警状态:', u'status:'], ssList=msg_list)
            rdic['level'] = self.search_from_list(seList=[u'告警等级:', u'severity:'], ssList=msg_list)
            rdic['info'] = self.search_from_list(seList=[u'告警信息:', u'trigger:'], ssList=msg_list)
            rdic['msg'] = self.search_from_list(seList=[u'问题详情:', u'item:'], ssList=msg_list)
            rdic['val'] = self.search_from_list(seList=[u'当前值：', u'当前值:', u'value:'], ssList=msg_list)
            if (rdic['state'] == 'OK'):
                msg = '---> recovery:\n' + msg #+ '\n' + self.__WT_SPLIT_STRING
            if len(msg) > self.ALERT_MAX_SIZE:
                msg = msg[0:self.ALERT_MAX_SIZE]
            return msg, rdic
        except Exception as e:
            return '', {}

    def alert_message_stat(self, alert_list=[], zbx_lvl=[]):
        rlm = []
        rld = []
        for alert in alert_list:
            msg, rdic = self.alert_message_init(alert=alert)
            if rdic:
                rld.append(rdic)
            if msg:
                for ss in zbx_lvl:
                    if ss in msg:
                        # 将信息添加到返回列表
                        rlm.append(msg)
        return rlm, rld

    def get_trigger_msg(self, zapi=ZabbixApi.ZabbixApi, zbx_name='', priority='0'):
        if (not zapi) or (not zapi.isLogin()):
            return ''
        try:
            para_ext = {
                'only_true': 1,
                'skipDependent': 1,
                'monitored': 1,
                'active': 1,
                'expandDescription': 1,
                'withLastEventUnacknowledged': 1
            }
            unack_triggers = zapi.trigger_get(output='description, value, priority, triggerid', selectGroups='group, name', selectHosts='host, name', para_ext=para_ext)
            msg = ''
            msgt = ''
            num = 0
            cdic = {}
            for t in unack_triggers:
                if (t['value'] == '1') and (t['priority'] >= priority):
                    num += 1
                    tag_msg = ''
                    tag_list = zapi.problem_get(objectids=t['triggerid'])
                    tnum = 0
                    for tag_m in tag_list:
                        if tag_m['tags']:
                            for tag_n in tag_m['tags']:
                                tnum += 1
                                tag_msg = u'%s\n    (%d)%s；' % (tag_msg, tnum, tag_n['value'])
                    if tnum > 0:
                        tag_msg = u'%d个tags：%s' % (tnum, tag_msg)
                    msgt = u'%s%d、%s 【%s】 %s [ %s ]\n' % (msgt, num, t['hosts'][0]['name'], t['priority'], t['description'], tag_msg)
                    g_name = t['groups'][0]['name']
                    if g_name in cdic:
                        cdic[g_name] += 1
                    else:
                        cdic[g_name] = 1
            ts = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(time.time()))
            msg = u'::: from %s 当前 :::\n%s\n%s\n%s\n当前告警 %d 个：\n' % (zbx_name, self.WT_SPLIT_STRING, ts, self.WT_SPLIT_STRING, num)
            for key in cdic:
                msg = u'%s%s: %d 个。\n' % (msg, key, cdic[key])
            return u'%s%s\n%s%s' % (msg, self.WT_SPLIT_STRING, msgt, self.WT_SPLIT_STRING)
        except Exception as e:
            print(e)
            return ''

    def get_group_from_list(self, hostid='', wsList=[]):
        try:
            for dic in wsList:
                if hostid == dic['hostid']:
                    return dic['group']
        except Exception as e:
            print(e)
        return ''

    def get_statistical_data(self, wsList=[]):
        rdic = {}
        for dic in wsList:
            try:
                if dic['state'] == 'OK':
                    continue
                if dic['group'] not in rdic:
                    rdic[dic['group']] = {}
                if dic['host'] not in rdic[dic['group']]:
                    rdic[dic['group']][dic['host']] = {}
                if dic['level'] in rdic[dic['group']][dic['host']]:
                    rdic[dic['group']][dic['host']][dic['level']] += 1
                else:
                    rdic[dic['group']][dic['host']][dic['level']] = 1
            except Exception as e:
                continue
        return rdic

    def get_statistical_msg(self, wsList=[]):
        msg = ''
        tdic = {}
        rdic = self.get_statistical_data(wsList=wsList)
        for group in rdic:
            tdic[group] = 0
            msg = u'%s%s:\n' % (msg, group)
            for host in rdic[group]:
                msg = u'%s    %s:\n' % (msg, host)
                for level in rdic[group][host]:
                    tdic[group] = tdic[group] + rdic[group][host][level]
                    msg = u'%s        %s: %d\n' % ( msg, level, rdic[group][host][level])
        if msg and tdic:
            num = 0
            msgt = ''
            for group in tdic:
                num += tdic[group]
                msgt = u'%s%s：%d 个\n' % (msgt, group, tdic[group])
            msg = u'累计告警：%d 个。\n%s%s\n%s%s' % (num, msgt, self.WT_SPLIT_STRING, msg, self.WT_SPLIT_STRING)
        return msg

    def save_statistical_file(self, wsList=[], data_file=''):
        if not wsList:
            return False
        if not data_file:
            data_file = self.data_file
        # 先删除文件
        if os.path.exists(data_file):
            os.remove(data_file)
        # 生成 xlsx 文件
        wb = Workbook(write_only=True)
        ws = []
        ws.append(wb.create_sheet(u'alert'))
        ws[0].append([u'主机群', u'主机', u'IP地址', u'时间', u'状态', u'等级', u'信息', u'详情', u'当前值'])
        for dic in wsList:
            ws[0].append([dic['group'], dic['host'], dic['IP'], dic['time'], dic['state'], dic['level'], dic['info'], dic['msg'], dic['val']])
        wb.save(data_file)
        wb.close()
        return True

    def send_alert_weixin(self, zbx_Dic={}, wx_Dic={}, startTime=0, endTime=0):
        try:
            zapi = zbx_Dic['zapi']
            wapi_list = [wx_Dic[zx]['wapi'] for zx in zbx_Dic['toagent'] if (zx in wx_Dic) and (wx_Dic[zx]['start_flag'])]
            if (not zbx_Dic['start_flag']) or (not zapi.isLogin()) or (not wapi_list):
                return
            # 获取告警信息
            alert_list = zapi.alert_get(output='alertid, message', selectHosts='host, name, hostid', time_from=startTime, time_till=endTime)
            # 告警消息进数据表
            alert_tm_list, alert_ws_list = self.alert_message_stat(alert_list=alert_list, zbx_lvl=zbx_Dic['alert_level'])
            if alert_ws_list:
                zbx_Dic['alert_ws_list'] = zbx_Dic['alert_ws_list'] + alert_ws_list
            if alert_tm_list:
                msg = ''
                for msgx in alert_tm_list:
                    msg = u'%s%s\n%s\n' % (msg, msgx, self.WT_SPLIT_STRING)
                msg = u'%s%s\n%s' %(msg, self.WT_SPLIT_BLOCK, self.get_trigger_msg(zapi=zapi, zbx_name=zbx_Dic['zbx_name'], priority=zbx_Dic['priority']))
                if len(msg) > self.WT_MSG_MAX_SIZE:
                    msg = msg[0:self.WT_MSG_MAX_SIZE]
                for wapi in wapi_list:
                    wapi.send_message(msg=msg)
        except Exception as e:
            print(e)

    def send_collect_weixin(self, zbx_Dic={}, wapi=WeTalkApi.WeTalkApi):
        try:
            zapi = zbx_Dic['zapi']
            if (not zbx_Dic['start_flag']) or (not zapi.isLogin()) or (not wapi.get_access_token()):
                return
            msg = self.get_trigger_msg(zapi=zapi, zbx_name=zbx_Dic['zbx_name'], priority=zbx_Dic['priority'])
            if msg:
                wapi.send_message(msg=msg)
        except Exception as e:
            print(e)

    def GetCollectInfoByList(self, zapi=ZabbixApi.ZabbixApi, zbx_name='', time_from=0, time_till=0, out_file=False, alert_ws_list=[]):
        if (not zapi) or (not zapi.isLogin()) or (not alert_ws_list):
            return '', ''
        try:
            for alert_dic in alert_ws_list:
                if alert_dic['group']:
                    continue
                # 查找主机的group_name
                if alert_dic['hostid']:
                    group_name = self.get_group_from_list(hostid=alert_dic['hostid'], wsList=alert_ws_list)
                    #print alert_dic['host'], 'find group_name = ', group_name
                    if not group_name:
                        try:
                            group_name = zapi.hostgroup_get(output='name', hostids=alert_dic['hostid'])[0]['name']
                        except Exception as e:
                            group_name = 'Unknown'
                    alert_dic['group'] = group_name
                else:
                    alert_dic['group'] = 'Unknown'
            # 初始化输出字符串
            msg = ''
            filename = ''
            t1 = time.strftime('%Y-%m-%d %H:%M:%S',time.localtime(time_from))
            t2 = time.strftime('%Y-%m-%d %H:%M:%S',time.localtime(time_till))
            msg = u'::: from %s :::\n%s\n%s -- %s\n%s\n' % (zbx_name, self.WT_SPLIT_STRING, t1, t2, self.WT_SPLIT_STRING)
            if alert_ws_list:
                msg = msg + self.get_statistical_msg(wsList=alert_ws_list)
                # 将数据写xls文件，发型文件到企业微信
                if out_file:
                    time_str = time.strftime('%Y%m%d%H%M%S',time.localtime(time.time()))
                    filename = u'Alert_detail_%s_%s.xlsx' % (zbx_name, time_str)
                    if not self.save_statistical_file(wsList=alert_ws_list, data_file=filename):
                        filename = ''
            return msg, filename
        except Exception as e:
            print(e)
            return '', ''

    def send_file_weixin(self, zbx_Dic={}, wapi=WeTalkApi.WeTalkApi, startTime=0, endTime=0):
        try:
            zapi = zbx_Dic['zapi']
            alert_ws_list = zbx_Dic['alert_ws_list']
            if (not zapi) or (not zapi.isLogin()) or (not wapi.get_access_token()) or (not alert_ws_list):
                return
            rmsg, filename = self.GetCollectInfoByList(zapi=zapi, zbx_name=zbx_Dic['zbx_name'], time_from=startTime, time_till=endTime, alert_ws_list=alert_ws_list)
            if rmsg:
                wapi.send_message(toparty=zbx_Dic['toparty'], msg=rmsg)
            if filename:
                wapi.send_message(msgtype='file', filename=filename)
        except Exception as e:
            print(e)

    def ShowAllCommand(self):
        if not self.Access_Command_dic:
            return ''
        num = 0
        msg = ''
        explain = u'命令格式为: 命令 参数1 参数2 ..., [ ]中为可选参数\ntn: 时间间隔，整数类型\nt1: 时间，如20180314或0314(默认当前年份)\nzbx: zabbix服务器, 如167,zbx_167\ngrp : 组名，多个以|进行分隔，如itms+|wap|dns\nf: 输出文件\n例：.getday 1 167 wap|dns f\n'
        for x in self.Access_Command_dic:
            num += 1
            msg = u'%s%d、%s：%s\n' % (msg, num, x, self.Access_Command_dic[x])
        msg = u'当前共有命令 %d 个：\n%s\n%s\n%s%s' % (num, self.WT_SPLIT_STRING, explain, msg, self.WT_SPLIT_STRING)
        zbx_list = []
        for zbx in self.z_dic:
            if self.z_dic[zbx]['start_flag']:
                zbx_list.append(zbx)
        msg = u'%s\nZBX:%s' % (msg, str(zbx_list))
        return msg

    def GetNow_AlertInfo(self, argstr=''):
        msg = ''
        zba = ''
        #arg_list = [x for x in argstr.split(' ') if x]
        arg_list = argstr.split()
        if arg_list:
            zba = arg_list[0]
        zba_list = [zbx for zbx in self.z_dic if (zba in zbx) and (self.z_dic[zbx]['start_flag'])]
        for zbx in zba_list:
            zapi = self.z_dic[zbx]['zapi']
            zbx_name = self.z_dic[zbx]['zbx_name']
            priority = self.z_dic[zbx]['priority']
            msg = u'%s%s\n%s\n' % (msg, self.get_trigger_msg(zapi=zapi, zbx_name=zbx_name, priority=priority), self.WT_SPLIT_BLOCK)
        return msg

    def GetCollect_AlertInfo(self, zapi=ZabbixApi.ZabbixApi, zbx_name='', groupname='', time_from=0, time_till=0, out_file=False):
        try:
            if (time_till <= time_from) or (not zapi) or (not zapi.isLogin()):
                return '', ''
            #time_till = int(time.time())
            #time_from = time_till - collect_time
            # 获取告警信息
            group_id_list = zapi.hostgroup_id_list(name=groupname)
            alert_list = zapi.alert_get(output='alertid, message', groupids=group_id_list, selectHosts='host, name, hostid', time_from=time_from, time_till=time_till)
            alert_ws_list = []
            for alert in alert_list:
                ss, rdic = self.alert_message_init(alert=alert)
                if rdic:
                    alert_ws_list.append(rdic)
            return self.GetCollectInfoByList(zapi=zapi, zbx_name=zbx_name, time_from=time_from, time_till=time_till, out_file=out_file, alert_ws_list=alert_ws_list)
        except Exception as e:
            print(e)
            return '', ''

    def GetTimeByStr(self, str_time=''):
        try:
            if (len(str_time) >= 2):
                td = {'s': 1, 'm': 60, 'h': 3600, 'd': 86400}
                smhd = str_time[-1:].lower()
                if smhd in td:
                    return int(float(eval(str_time[:-1])) * td[smhd])
            return -1
        except Exception as e:
            return -1

    def GetTimeStampByStr(self, str_time='', year_num=10):
        try:
            if not str_time.isdigit():
                return -1
            now_year = time.localtime()[0]
            if len(str_time) < 4:
                str_time = str(now_year) + str_time
            else:
                now_year_list = []
                for i in range(0, year_num + 1):
                    now_year_list.append(str(now_year - i))
                if str_time[0:4] not in now_year_list:
                    str_time = str(now_year) + str_time
            s = str_time.ljust(14, '0')
            timestring = '%s-%s-%s %s:%s:%s' % (s[0:4], s[4:6], s[6:8], s[8:10], s[10:12], s[12:14])
            return int(time.mktime(time.strptime(timestring, '%Y-%m-%d %H:%M:%S')))
        except Exception as e:
            return -1

    def Gettt_AlertInfo(self, argstr=''):
        try:
            #arg_list = [x for x in argstr.split(' ') if x]
            arg_list = argstr.split()
            argnum = len(arg_list)
            if argnum < 2:
                return u'错误：参数错误！', ''
            # 解析输入参数
            time_from = self.GetTimeStampByStr(arg_list[0])
            time_till = self.GetTimeStampByStr(arg_list[1])
            if (time_till <= 0):
                time_till = time_from + self.GetTimeByStr(arg_list[1])
            if (time_from <= 0) or (time_till <= 0) or (time_till <= time_from):
                return u'错误：时间参数错误！', ''
            zba = ''
            groupname = ''
            out_file = False
            if argnum >= 3:
                if arg_list[2] == 'f':
                    out_file = True
                else:
                    zba = arg_list[2]
                    if argnum >= 4:
                        if arg_list[3] == 'f':
                            out_file = True
                        else:
                            groupname = arg_list[3]
                            if (argnum >= 5) and (arg_list[4] == 'f'):
                                out_file = True
            zba_list = [zbx for zbx in self.z_dic if (zba in zbx) and (self.z_dic[zbx]['start_flag'])]
            if not zba_list:
                return u'错误：zbx名错误或未启动！', ''
            # 获得匹配的第一个zbx server的统计数据
            zapi = self.z_dic[zba_list[0]]['zapi']
            zbx_name = self.z_dic[zba_list[0]]['zbx_name']
            return self.GetCollect_AlertInfo(zapi=zapi, zbx_name=zbx_name, groupname=groupname, time_from=time_from, time_till=time_till, out_file=out_file)
        except Exception as e:
            return u'错误：参数错误！', ''

    def Get_AlertInfo(self, argstr='', bs=1):
        try:
            #arg_list = [x for x in argstr.split(' ') if x]
            arg_list = argstr.split()
            argnum = len(arg_list)
            if argnum < 1:
                return u'错误：参数错误！', ''
            # 解析输入参数
            collect_time = int(eval(arg_list[0])) * bs
            zba = ''
            groupname = ''
            out_file = False
            if argnum >= 2:
                if arg_list[1] == 'f':
                    out_file = True
                else:
                    zba = arg_list[1]
                    if argnum >= 3:
                        if arg_list[2] == 'f':
                            out_file = True
                        else:
                            groupname = arg_list[2]
                            if (argnum >= 4) and (arg_list[3] == 'f'):
                                out_file = True
            zba_list = [zbx for zbx in self.z_dic if (zba in zbx) and (self.z_dic[zbx]['start_flag'])]
            if not zba_list:
                return u'错误：zbx名错误或未启动！', ''
            # 获得匹配的第一个zbx server的统计数据
            zapi = self.z_dic[zba_list[0]]['zapi']
            zbx_name = self.z_dic[zba_list[0]]['zbx_name']
            time_till = int(time.time())
            time_from = time_till - collect_time
            return self.GetCollect_AlertInfo(zapi=zapi, zbx_name=zbx_name, groupname=groupname, time_from=time_from, time_till=time_till, out_file=out_file)
        except Exception as e:
            return u'错误：参数错误！', ''

    def AutoGetAlertInfo(self, cmd_msg=''):
        # 对消息进行处理
        cmd_list = [x for x in self.Access_Command_dic if x == cmd_msg[0:len(x)]]
        if not cmd_list:
            return '', ''
        # 取第一个操作指令
        cmd = cmd_list[0]
        if cmd == '.help':
            return self.ShowAllCommand(), ''
        elif cmd == '.getnow':
            argstr = cmd_msg[len(cmd):].strip()
            return self.GetNow_AlertInfo(argstr=argstr), ''
        elif cmd == '.gettt':
            argstr = cmd_msg[len(cmd):].strip()
            return self.Gettt_AlertInfo(argstr=argstr)
        elif cmd in self.Time_BS_dic:
            argstr = cmd_msg[len(cmd):].strip()
            return self.Get_AlertInfo(argstr=argstr, bs=self.Time_BS_dic[cmd])

    def AutoSendAlertInfo(self, wapi=WeTalkApi.WeTalkApi, cmd_msg='', touser=''):
        if (not cmd_msg) or (not wapi) or (not wapi.get_access_token()):
            return ''
        # 获得查询命令返回的信息
        rmsg, filename = self.AutoGetAlertInfo(cmd_msg=cmd_msg)
        if rmsg:
            date_time = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(time.time()))
            rmsg = '[%s]: %s\n%s\n%s' % (date_time, cmd_msg, self.WT_SPLIT_STRING, rmsg)
            wapi.send_message(msg=rmsg, touser=touser)
        if filename:
            wapi.send_message(msgtype='file', filename=filename, touser=touser)

    def run(self):
        # 变量赋值
        z_dic = self.z_dic
        w_dic = self.w_dic
        # 判断zabbix和企业微信是否OK？
        if (not [zbx for zbx in z_dic if z_dic[zbx]['start_flag']]):
            print 'error: all zabbix stop, exit!'
            return
        if (not [wx for wx in w_dic if w_dic[wx]['start_flag']]):
            print 'error: all wexin agent stop, exit!'
            return

        # 启动程序主程序，可用循环语句，也可用线程实现。
        while True:
            time.sleep(1)
            now_time = int(time.time())
            for zbx in z_dic:
                if not z_dic[zbx]['start_flag']:
                    continue
                if (now_time - z_dic[zbx]['alert_last_time']) >= z_dic[zbx]['alert_time']:
                    # 发送告警信息到企业微信
                    self.send_alert_weixin(zbx_Dic=z_dic[zbx], wx_Dic=w_dic, startTime=z_dic[zbx]['alert_last_time'], endTime=now_time)
                    z_dic[zbx]['alert_last_time'] = now_time
                if (now_time - z_dic[zbx]['collect_last_time']) >= z_dic[zbx]['collect_time']:
                    # 发送告警汇总信息到企业微信
                    self.send_collect_weixin(zbx_Dic=z_dic[zbx], wapi=w_dic[self.wt_file_sec]['wapi'])
                    z_dic[zbx]['collect_last_time'] = now_time
                if (now_time - z_dic[zbx]['reset_last_time']) >= z_dic[zbx]['reset_time']:
                    # 发送统计信息和详单文件到企业微信
                    self.send_file_weixin(zbx_Dic=z_dic[zbx], wapi=w_dic[self.wt_file_sec]['wapi'], startTime=z_dic[zbx]['reset_last_time'], endTime=now_time)
                    # 每天清空一次详单列表
                    del z_dic[zbx]['alert_ws_list'][:]
                    z_dic[zbx]['reset_last_time'] = now_time
                    z_dic[zbx]['info_last_time'] = now_time
                if (now_time - z_dic[zbx]['info_last_time']) >= z_dic[zbx]['info_time']:
                    # 发送统计信息和详单文件到企业微信
                    self.send_file_weixin(zbx_Dic=z_dic[zbx], wapi=w_dic[self.wt_file_sec]['wapi'], startTime=z_dic[zbx]['reset_last_time'], endTime=now_time)
                    z_dic[zbx]['info_last_time'] = now_time

    def WeTalkMenu(self):
        z_dic = self.z_dic
        w_dic = self.w_dic
        menu = {
            'button': [
                {'type': 'click', 'name': u'命令帮助', 'key': '.help'},
                {'name': u'当前告警', 'sub_button': [{'type': 'click', 'name': u'所有', 'key': '.getnow'}]},
                {'name': u'昨天告警', 'sub_button': []},
            ]
        }
        for zbx in z_dic:
             if z_dic[zbx]['start_flag'] != 0:
                 m_dic = {'type': 'click', 'name': z_dic[zbx]['zbx_name'], 'key': '.getnow ' + z_dic[zbx]['zbx_name']}
                 menu['button'][1]["sub_button"].append(m_dic)
                 m_dic = {'type': 'click', 'name': z_dic[zbx]['zbx_name'], 'key': '.getday 1 ' + z_dic[zbx]['zbx_name'] + ' f'}
                 menu['button'][2]["sub_button"].append(m_dic)
        wapi = w_dic[self.wt_iGet_sec]['wapi']
        if wapi:
            wapi.send_menu(method='delete')
            wapi.send_menu(method='create', menu=menu)
## ZabbixWeTalkApi class end -------------------------------------------------------------------------------------------------

